#! /usr/bin/env python
# -*- coding: utf-8 -*-

'''doc'''




from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers, PatternFill, Side, Border


from .tool import to_str, to_float, Data



class Cell():
    '''doc'''
    def __init__(self, cell):
        self.cell = cell

    def get_val(self):
        '''doc'''
        return self.cell.value

    def get_float_val(self):
        '''doc'''
        val = self.get_val()
        return to_float(val)

    def get_str_val(self):
        '''doc'''
        val = self.get_val()
        return to_str(val)

    def set_val(self, val):
        '''doc'''
        self.cell.value = val

    def set_number_format_text(self):
        '''doc'''
        self.cell.number_format = numbers.FORMAT_TEXT

    def set_fill_red(self):
        '''doc'''
        self.cell.fill = PatternFill(fill_type="solid", \
                            start_color="FFCCFF", end_color="FFCCFF")

    def set_fill_blue(self):
        '''doc'''
        self.cell.fill = PatternFill(fill_type="solid", \
                            start_color="CCFFFF", end_color="CCFFFF")

    def set_border_thin(self):
        '''doc'''
        thin_side = Side(border_style="thin", color="000000")
        thin_border = Border(top=thin_side, left=thin_side, \
                            right=thin_side, bottom=thin_side)
        self.cell.border = thin_border

class Sheet():
    '''doc'''
    def __init__(self, sheet):
        self.sheet = sheet

        self.num_col_index = None
        self.num_col_data = None
        self.alone_val_rows = None
        self.sheet_cell_dict = {}

    def get_cell(self, row, col):
        '''doc'''
        sheet_cell = self.sheet_cell_dict.get((row, col))
        if sheet_cell is not None:
            return sheet_cell
        sheet_cell = self.sheet.cell(row=row, column=col)
        cell = Cell(sheet_cell)
        self.sheet_cell_dict[row, col] = cell
        return cell
    def exist_row(self, src_col_vals):
        '''doc'''
        max_row = self.get_max_row()
        max_col = self.get_max_col()
        src_col_num = len(src_col_vals)
        for row_index in range(1, max_row + 1):
            all_equal = True
            for col_index in range(1, max_col + 1):
                if src_col_num < col_index:
                    break
                src_val = src_col_vals[col_index - 1]
                tar_val = self.get_cell(row_index, col_index).get_str_val()


    def get_max_col(self):
        '''doc'''
        sheet_max_row = self.sheet.max_row
        sheet_max_col = self.sheet.max_column
        max_col = sheet_max_col
        for col_index in range(sheet_max_col, 0, -1):
            has_val = False
            for row_index in range(sheet_max_row, 0, -1):
                val = self.get_cell(row_index, col_index).get_str_val()
                if val is not None:
                    has_val = True
                    break
            if not has_val:
                max_col -= 1
            else:
                break
        return max_col

    def get_max_row(self):
        '''doc'''
        sheet_max_row = self.sheet.max_row
        sheet_max_col = self.sheet.max_column
        max_row = sheet_max_row
        for row_index in range(sheet_max_row, 0, -1):
            has_val = False
            for col_index in range(sheet_max_col, 0, -1):
                val = self.get_cell(row_index, col_index).get_str_val()
                if val is not None:
                    has_val = True
                    break
            if not has_val:
                max_row -= 1
            else:
                break
        return max_row

    def get_num_col_index(self):
        '''doc'''
        if self.num_col_index is not None:
            return self.num_col_index

        max_col = self.get_max_col()
        max_row = self.get_max_row()

        num_col_count = 0
        num_col_index = 0

        for col_index in range(1, max_col + 1):
            num_cell_count = 0
            is_str_cell = False
            for row_index in range(1, max_row + 1):
                cell = self.get_cell(row_index, col_index)
                float_val = cell.get_float_val()

                if isinstance(float_val, float):
                    num_cell_count += 1
                    continue

                str_val = cell.get_str_val()

                if isinstance(str_val, str):
                    is_str_cell = True
                    break

            if is_str_cell:
                continue

            if num_cell_count >= 1:
                num_col_count += 1
                num_col_index = col_index

        data = Data()

        if num_col_count == 0:
            data.set("ERR", True)
            data.set("ERR_MSG", "没有数字列")
            return None

        if num_col_count > 1:
            data.set("ERR", True)
            data.set("ERR_MSG", "有%d列是数字" % num_col_count)
            return None

        if num_col_count == 1:
            self.num_col_index = num_col_index

        return self.num_col_index

    def get_num_col_data(self):
        '''doc'''
        if self.num_col_data is not None:
            return self.num_col_data

        num_col_index = self.num_col_index

        val_cnt_dict = {}
        val_set = set()
        val_list = []
        row_val_dict = {}
        row_list = []

        max_row = self.get_max_row()


        for row_index in range(1, max_row + 1):
            cell = self.get_cell(row_index, num_col_index)
            float_val = cell.get_float_val()
            if isinstance(float_val, float):
                val_set.add(float_val)
                val_list.append(float_val)
                row_val_dict[row_index] = float_val
                row_list.append(row_index)

                val_cnt = val_cnt_dict.get(float_val)
                if val_cnt is None:
                    val_cnt_dict[float_val] = 1
                else:
                    val_cnt_dict[float_val] = val_cnt + 1

        self.num_col_data = {}
        self.num_col_data["VAL_SET"] = val_set
        self.num_col_data["VAL_LIST"] = val_list
        self.num_col_data["ROW_VAL_DICT"] = row_val_dict
        self.num_col_data["ROW_LIST"] = row_list
        self.num_col_data["VAL_CNT_DICT"] = val_cnt_dict
        return self.num_col_data

    def get_alone_val_rows(self):
        '''doc'''
        if self.alone_val_rows is not None:
            return self.alone_val_rows
        row_list = self.num_col_data.get("ROW_LIST")
        row_val_dict = self.num_col_data.get("ROW_VAL_DICT")
        val_cnt_dict = self.num_col_data.get("VAL_CNT_DICT")

        self.alone_val_rows = []
        for row_index in row_list:
            val = row_val_dict.get(row_index)
            count = val_cnt_dict.get(val)
            if count == 1:
                self.alone_val_rows.append(row_index)
        return self.alone_val_rows

    def get_rows_by_val(self, val):
        '''doc'''
        val = to_float(val)
        ret = []
        row_list = self.num_col_data.get("ROW_LIST")
        row_val_dict = self.num_col_data.get("ROW_VAL_DICT")
        for row_index in row_list:
            find_val = row_val_dict.get(row_index)
            if find_val == val:
                ret.append(row_index)
        return ret

    def copy_row_from_sheet(self, src_sheet, row_index, color):
        '''doc'''

        tar_max_row = self.get_max_row()
        src_max_col = src_sheet.get_max_col()

        for src_col_index in range(1, src_max_col + 1):
            src_cell = src_sheet.get_cell(row_index, src_col_index)
            tar_cell = self.get_cell(tar_max_row + 1, src_col_index)
            tar_cell.set_val(src_cell.get_val())

            tar_cell.set_border_thin()

            tar_cell.set_number_format_text()

            if color == "RED":
                tar_cell.set_fill_red()
            elif color == "BLUE":
                tar_cell.set_fill_blue()

class Book():
    '''doc'''
    def __init__(self, path):
        self.path = path
        self.book = None
        self.book_sheet_dict = {}

    def get_active_sheet(self):
        '''doc'''
        active_sheet = self.book.get_active_sheet()
        return Sheet(active_sheet)

    def create(self):
        '''doc'''
        self.book = Workbook(write_only=False)

    def load(self):
        '''doc'''
        self.book = load_workbook(self.path, read_only=False, keep_vba=False, \
            data_only=True, guess_types=False, keep_links=False)

    def create_sheet(self, title=None, index=None):
        '''doc'''
        self.book.create_sheet(title, index)

    def get_sheet(self, sheet_name):
        '''doc'''
        if self.book_sheet_dict.get(sheet_name) is not None:
            return self.book_sheet_dict.get(sheet_name)
        book_sheet = self.book.get_sheet_by_name(sheet_name)
        sheet = Sheet(book_sheet)
        self.book_sheet_dict[sheet_name] = sheet
        return sheet

    def get_sheet_names(self):
        '''doc'''
        return self.book.get_sheet_names()

    def has_sheet(self, *sheet_names):
        '''doc'''
        sheet_names_len = len(sheet_names)
        if sheet_names_len == 0:
            return False
        book_sheet_names = self.get_sheet_names()
        for sheet_name in sheet_names:
            sheet_name = to_str(sheet_name)
            if sheet_name not in book_sheet_names:
                return False
        return True

    def save(self):
        '''doc'''
        self.book.save(self.path)

    def close(self):
        '''doc'''
        if self.book is not None:
            self.book.close()
        self.book = None
