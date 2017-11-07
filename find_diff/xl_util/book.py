#! /usr/bin/env python
# -*- coding: utf-8 -*-

'''doc'''




from openpyxl import load_workbook, Workbook

from util import to_str
from sheet import Sheet

class Book():
    '''doc'''
    def __init__(self, path):
        self.path = path
        self.book = None
        self.book_sheet_dict = {}

    def get_active_sheet(self):
        '''doc'''
        active_sheet = self.book.get_active_sheet()
        return Sheet(active_sheet)

    def create(self):
        '''doc'''
        self.book = Workbook(write_only=False)

    def load(self):
        '''doc'''
        self.book = load_workbook(self.path, read_only=False, keep_vba=False, \
            data_only=True, guess_types=False, keep_links=False)

    def create_sheet(self, title=None, index=None):
        '''doc'''
        self.book.create_sheet(title, index)

    def get_sheet(self, sheet_name):
        '''doc'''
        if self.book_sheet_dict.get(sheet_name) is not None:
            return self.book_sheet_dict.get(sheet_name)
        book_sheet = self.book.get_sheet_by_name(sheet_name)
        sheet = Sheet(book_sheet)
        self.book_sheet_dict[sheet_name] = sheet
        return sheet

    def get_sheet_names(self):
        '''doc'''
        return self.book.get_sheet_names()

    def has_sheet(self, *sheet_names):
        '''doc'''
        sheet_names_len = len(sheet_names)
        if sheet_names_len == 0:
            return False
        book_sheet_names = self.get_sheet_names()
        for sheet_name in sheet_names:
            sheet_name = to_str(sheet_name)
            if sheet_name not in book_sheet_names:
                return False
        return True

    def save(self):
        '''doc'''
        self.book.save(self.path)

    def close(self):
        '''doc'''
        if self.book is not None:
            self.book.close()
        self.book = None
