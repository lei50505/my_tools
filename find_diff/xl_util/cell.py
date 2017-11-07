#! /usr/bin/env python
# -*- coding: utf-8 -*-

'''doc'''





from openpyxl.styles import numbers, PatternFill, Side, Border


from util import to_str, to_float



class Cell():
    '''doc'''
    def __init__(self, cell):
        self.cell = cell

    def get_val(self):
        '''doc'''
        return self.cell.value

    def get_float_val(self):
        '''doc'''
        val = self.get_val()
        return to_float(val)

    def get_str_val(self):
        '''doc'''
        val = self.get_val()
        return to_str(val)

    def set_val(self, val):
        '''doc'''
        self.cell.value = val

    def set_number_format_text(self):
        '''doc'''
        self.cell.number_format = numbers.FORMAT_TEXT

    def set_fill_red(self):
        '''doc'''
        self.cell.fill = PatternFill(fill_type="solid", \
                            start_color="FFCCFF", end_color="FFCCFF")

    def set_fill_blue(self):
        '''doc'''
        self.cell.fill = PatternFill(fill_type="solid", \
                            start_color="CCFFFF", end_color="CCFFFF")

    def set_border_thin(self):
        '''doc'''
        thin_side = Side(border_style="thin", color="000000")
        thin_border = Border(top=thin_side, left=thin_side, \
                            right=thin_side, bottom=thin_side)
        self.cell.border = thin_border
